#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Сводный файл по конкурентам KNRU / Nordwest / Rest2Rent.

Логика:
- объединяем один и тот же объект по адресу (street+house+corp/str+deal_type)
  и близкой площади (±3 м2);
- показываем 3 колонки присутствия: KNRU / Северо-Запад / Rest2Rent;
- если значения различаются между конкурентами, пишем это в колонке "Расхождения";
- отдельный лист "Статистика" с пересечениями и количеством расхождений.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Dict, List, Optional
import urllib.parse
import urllib.request
import ssl
import time

import pandas as pd
from openpyxl.styles import PatternFill

import robot


BASE_DIR = Path(__file__).resolve().parent
TODAY = str(date.today())
AREA_TOL = 3.0


@dataclass
class Listing:
    competitor: str
    competitor_title: str
    deal_type: str
    address: str
    area_m2: Optional[float]
    price_rub: Optional[float]
    result: str
    price_alert: str
    position_global: Optional[float]
    competitor_link: str
    our_link: str
    our_best_price: Optional[float]
    district: str
    district_norm: str
    street_key: str
    address_key: str


@dataclass
class UnionObject:
    address_key: str
    area_ref: Optional[float] = None
    listings: Dict[str, Listing] = field(default_factory=dict)
    area_values: List[float] = field(default_factory=list)

    def add_listing(self, lst: Listing):
        cur = self.listings.get(lst.competitor)
        if cur is None:
            self.listings[lst.competitor] = lst
        else:
            cur_pos = cur.position_global if cur.position_global is not None else 10**12
            new_pos = lst.position_global if lst.position_global is not None else 10**12
            if new_pos < cur_pos:
                self.listings[lst.competitor] = lst

        if isinstance(lst.area_m2, (int, float)):
            self.area_values.append(float(lst.area_m2))
            self.area_ref = float(lst.area_m2)


def parse_num(value) -> Optional[float]:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    s = s.replace(" ", "").replace(",", ".")
    s = re.sub(r"[^\d\.-]", "", s)
    if not s:
        return None
    try:
        return float(s)
    except Exception:
        return None


def format_money(v: Optional[float]) -> str:
    if not isinstance(v, (int, float)):
        return ""
    return f"{int(round(float(v))):,}".replace(",", " ")


def format_area(v: Optional[float]) -> str:
    if not isinstance(v, (int, float)):
        return ""
    return f"{float(v):.1f}".rstrip("0").rstrip(".")


def to_abs_path(path_str: str) -> Path:
    p = Path(path_str)
    return p if p.is_absolute() else (BASE_DIR / p)


def build_address_key(address: str) -> str:
    comp = robot.extract_components(address)
    if comp and (comp.get("street_key_bag") or comp.get("street_key")):
        street = comp.get("street_key_bag") or comp.get("street_key") or ""
        hf = comp.get("house_from")
        ht = comp.get("house_to")
        corp = (comp.get("corp") or "").strip().lower()
        stro = (comp.get("str") or "").strip().lower()
        house = ""
        if hf is not None:
            house = f"{hf}-{ht or hf}"
        return f"{street}|{house}|{corp}|{stro}"

    # fallback, если адрес не разобрался
    norm = robot.norm_text(address or "")
    return f"fallback|{norm}"


def clean_city_prefix(address: str) -> str:
    s = (address or "").replace("\xa0", " ").strip()
    if not s:
        return s
    # Приводим разные тире к обычному дефису.
    s = re.sub(r"[‐‑‒–—−﹘﹣－]", "-", s)
    patterns = [
        r"^\s*(?:россия,\s*)?(?:г\.?\s*)?санкт(?:-|\s)?петербург(?:\s*г\.?)?\s*,\s*",
        r"^\s*(?:россия,\s*)?спб\s*,\s*",
    ]
    out = s
    # У некоторых источников префикс города дублируется дважды.
    for _ in range(3):
        prev = out
        for p in patterns:
            out = re.sub(p, "", out, flags=re.I)
        if out == prev:
            break
    return out.strip()


def normalize_district(value: str) -> str:
    s = (value or "").strip().lower()
    if not s:
        return ""
    s = robot.norm_text(s)
    s = s.replace("муниципальный", "")
    s = s.replace("район", "")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def is_unknown_district(norm_value: str) -> bool:
    s = (norm_value or "").strip().lower()
    return (not s) or s.startswith("не определ")


def infer_region_from_address(address: str) -> str:
    n = robot.norm_text(address or "")
    if not n:
        return ""

    # Явные области.
    if "новгородск" in n:
        return "Новгородская область"

    if "ленинградск" in n or "лен обл" in n:
        guessed = infer_lenoblast_raion(n)
        if guessed:
            return f"Ленинградская область, {guessed}"

        # Пытаемся вытащить конкретный район области.
        m = re.search(r"\b([а-яё\-]+)\s*(?:муниципальный\s*)?район\b", n)
        if m:
            part = m.group(1).strip()
            return f"Ленинградская область, {part.title()} район"

        return "Ленинградская область"

    # Иногда в адресе напрямую фигурирует "X район", но без области.
    m = re.search(r"\b([а-яё\-]+)\s*(?:муниципальный\s*)?район\b", n)
    if m and "санкт" not in n:
        part = m.group(1).strip()
        return f"{part.title()} район"

    return ""


def choose_top(counts: Dict[str, int]) -> str:
    if not counts:
        return ""
    return sorted(counts.items(), key=lambda x: (-x[1], x[0]))[0][0]


SPB_SUBAREA_TO_DISTRICT = {
    "рыбацкое": "Невский",
    "новая деревня": "Приморский",
    "пески": "Центральный",
    "округ новоизмайловское": "Московский",
    "округ измайловское": "Адмиралтейский",
    "екатерингофский округ": "Адмиралтейский",
    "округ чкаловское": "Петроградский",
    "округ академическое": "Калининский",
    "финляндский округ": "Калининский",
    "округ смольнинское": "Центральный",
    "округ коломна": "Адмиралтейский",
    "округ семеновский": "Адмиралтейский",
    "округ васильевский": "Василеостровский",
    "парголово": "Выборгский",
    "шушары": "Пушкинский",
    "горелово": "Красносельский",
}

LEN_SETTLEMENT_TO_RAION = {
    "кудрово": "Всеволожский район",
    "всеволожск": "Всеволожский район",
    "бугры": "Всеволожский район",
    "мурино": "Всеволожский район",
    "девяткино": "Всеволожский район",
    "колтуш": "Всеволожский район",
    "янино": "Всеволожский район",
    "разметелево": "Всеволожский район",
    "кальтино": "Всеволожский район",
    "мяглово": "Всеволожский район",
    "порошкино": "Всеволожский район",
    "кузьмолов": "Всеволожский район",
    "куйвози": "Всеволожский район",
    "имени морозова": "Всеволожский район",
    "токсово": "Всеволожский район",
    "мистолово": "Всеволожский район",
    "лупполово": "Всеволожский район",
    "скотное": "Всеволожский район",
    "хиттолово": "Всеволожский район",
    "нижние осельки": "Всеволожский район",
    "новосаратовка": "Всеволожский район",
    "сертолово": "Всеволожский район",
    "щеглово": "Всеволожский район",
    "проба": "Всеволожский район",
    "аннино": "Ломоносовский район",
    "новоселье": "Ломоносовский район",
    "виллози": "Ломоносовский район",
    "горбунки": "Ломоносовский район",
    "коваш": "Ломоносовский район",
    "порзолово": "Ломоносовский район",
    "пигелево": "Ломоносовский район",
    "новогорелово": "Ломоносовский район",
    "узигонты": "Ломоносовский район",
    "яльгелево": "Ломоносовский район",
    "санино": "Ломоносовский район",
    "гатчина": "Гатчинский район",
    "федоровск": "Тосненский район",
    "фёдоровск": "Тосненский район",
    "любан": "Тосненский район",
    "нурма": "Тосненский район",
    "тосно": "Тосненский район",
    "рябово": "Тосненский район",
    "волхов": "Волховский район",
    "войсковицы": "Гатчинский район",
    "елизаветино": "Гатчинский район",
    "малое верево": "Гатчинский район",
    "куровицы": "Гатчинский район",
    "рождествено": "Гатчинский район",
    "шпаньково": "Гатчинский район",
    "мины": "Гатчинский район",
    "волосово": "Волосовский район",
    "сланцы": "Сланцевский район",
    "шлиссельбург": "Кировский район",
    "мга": "Кировский район",
    "раздолье": "Приозерский район",
    "пески": "Выборгский район",
    "первомайское": "Выборгский район",
    "луга": "Лужский район",
}


def load_geo_cache(path: Path) -> Dict[str, str]:
    if not path.exists():
        return {}
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
        return raw if isinstance(raw, dict) else {}
    except Exception:
        return {}


def save_geo_cache(path: Path, data: Dict[str, str]):
    try:
        path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass


def _pick_geocode_candidate(addr: dict) -> str:
    state = (addr.get("state") or "").strip()
    county = (addr.get("county") or "").strip()
    state_district = (addr.get("state_district") or "").strip()
    city = (addr.get("city") or "").strip()
    town = (addr.get("town") or "").strip()
    village = (addr.get("village") or "").strip()
    municipality = (addr.get("municipality") or "").strip()
    city_district = (addr.get("city_district") or "").strip()
    suburb = (addr.get("suburb") or "").strip()
    borough = (addr.get("borough") or "").strip()

    state_norm = robot.norm_text(state)
    city_norm = robot.norm_text(city)

    if "ленинград" in state_norm:
        sub = county or state_district or municipality or city_district or town or village or suburb
        guessed = infer_lenoblast_raion(sub or "")
        if guessed:
            return f"Ленинградская область, {guessed}"
        if sub:
            return f"Ленинградская область, {sub}"
        return "Ленинградская область"

    if "новгород" in state_norm:
        sub = county or state_district or municipality or town or village
        if sub:
            return f"Новгородская область, {sub}"
        return "Новгородская область"

    if "санкт-петербург" in state_norm or "санкт-петербург" in city_norm:
        sub = city_district or suburb or municipality or borough or county
        if sub:
            sub_norm = normalize_district(sub)
            if sub_norm in SPB_SUBAREA_TO_DISTRICT:
                return SPB_SUBAREA_TO_DISTRICT[sub_norm]
            return sub
        return "Санкт-Петербург"

    if state:
        return state
    return ""


def infer_lenoblast_raion(text: str) -> str:
    n = robot.norm_text(text or "")
    for key, raion in LEN_SETTLEMENT_TO_RAION.items():
        if key in n:
            return raion
    return ""


def geocode_district(address: str, cache: Dict[str, str], timeout_sec: int = 8) -> str:
    q = (address or "").strip()
    if not q:
        return ""
    if q in cache:
        return cache[q]

    query = q
    # Для городских адресов добавляем СПб, если нет явной области.
    n = robot.norm_text(q)
    if ("обл" not in n and "область" not in n and "санкт" not in n and "спб" not in n):
        query = f"{q}, Санкт-Петербург"

    url = "https://nominatim.openstreetmap.org/search?" + urllib.parse.urlencode(
        {
            "q": query,
            "format": "jsonv2",
            "addressdetails": 1,
            "limit": 1,
            "accept-language": "ru",
        }
    )
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": "ParserDistrictBot/1.0 (local workspace)",
            "Accept": "application/json",
        },
    )

    try:
        ctx = ssl._create_unverified_context()
        with urllib.request.urlopen(req, timeout=timeout_sec, context=ctx) as resp:
            payload = json.loads(resp.read().decode("utf-8", "ignore"))
    except Exception:
        cache[q] = ""
        return ""

    if not payload:
        cache[q] = ""
        return ""

    addr = payload[0].get("address") or {}
    out = _pick_geocode_candidate(addr)
    cache[q] = out
    return out


def enrich_missing_districts(
    listings: List[Listing],
    use_geocode: bool = True,
    geocode_limit: int = 500,
    geocode_delay_sec: float = 1.1,
    geocode_timeout_sec: int = 8,
    geocode_cache_path: Path | None = None,
):
    cache_path = geocode_cache_path or (BASE_DIR / "district_cache_nominatim.json")
    geo_cache: Dict[str, str] = load_geo_cache(cache_path) if use_geocode else {}
    geocode_used = 0

    # Мапы знаний: точный адрес и улица -> район.
    addr_counts: Dict[str, Dict[str, int]] = {}
    street_counts: Dict[str, Dict[str, int]] = {}
    display_by_norm: Dict[str, str] = {}

    for x in listings:
        if not x.district_norm:
            continue
        display_by_norm.setdefault(x.district_norm, x.district or x.district_norm.title())
        addr_counts.setdefault(x.address_key, {})
        addr_counts[x.address_key][x.district_norm] = addr_counts[x.address_key].get(x.district_norm, 0) + 1
        if x.street_key:
            street_counts.setdefault(x.street_key, {})
            street_counts[x.street_key][x.district_norm] = street_counts[x.street_key].get(x.district_norm, 0) + 1

    for x in listings:
        if x.district_norm:
            continue

        chosen_norm = ""
        chosen_display = ""

        # 1) Точный ключ адреса.
        c1 = addr_counts.get(x.address_key, {})
        if c1:
            chosen_norm = choose_top(c1)
            chosen_display = display_by_norm.get(chosen_norm, chosen_norm.title())

        # 2) По улице, если район стабилен.
        if not chosen_norm and x.street_key:
            c2 = street_counts.get(x.street_key, {})
            if c2:
                total = sum(c2.values())
                top_norm, top_cnt = sorted(c2.items(), key=lambda z: (-z[1], z[0]))[0]
                if len(c2) == 1 and top_cnt >= 1:
                    chosen_norm = top_norm
                    chosen_display = display_by_norm.get(chosen_norm, chosen_norm.title())
                elif total > 0 and (top_cnt / total) >= 0.60 and top_cnt >= 2:
                    chosen_norm = top_norm
                    chosen_display = display_by_norm.get(chosen_norm, chosen_norm.title())

        # 3) Области/районы из текста адреса.
        if not chosen_norm:
            region_display = infer_region_from_address(x.address)
            if region_display:
                chosen_display = region_display
                chosen_norm = normalize_district(region_display)

        # 4) Nominatim для остатка.
        if not chosen_norm and use_geocode and geocode_used < geocode_limit:
            cached_before = x.address in geo_cache
            geo = geocode_district(x.address, geo_cache, timeout_sec=geocode_timeout_sec)
            if geo:
                chosen_display = geo
                chosen_norm = normalize_district(geo)
            if not cached_before:
                geocode_used += 1
            if (not cached_before) and geocode_delay_sec > 0:
                time.sleep(geocode_delay_sec)

        # 4b) Уточнение Ленобласти по тексту адреса.
        if chosen_display.startswith("Ленинградская область") and "," not in chosen_display:
            guessed = infer_lenoblast_raion(x.address)
            if guessed:
                chosen_display = f"Ленинградская область, {guessed}"
                chosen_norm = normalize_district(chosen_display)

        # 5) Фоллбек.
        if not chosen_display:
            chosen_display = "Не определен"
            chosen_norm = normalize_district(chosen_display)

        x.district = chosen_display
        x.district_norm = chosen_norm

    if use_geocode:
        save_geo_cache(cache_path, geo_cache)


def read_competitor_csv(path: Path, competitor: str, competitor_title: str, default_deal_type: str) -> List[Listing]:
    if not path.exists():
        return []

    out: List[Listing] = []
    with open(path, encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            address_raw = (row.get("address") or "").strip()
            address = clean_city_prefix(address_raw)
            if not address:
                continue

            district = (row.get("district") or "").strip()
            district_norm = normalize_district(district)
            comp = robot.extract_components(address)
            street_key = ""
            if comp:
                street_key = (comp.get("street_key_bag") or comp.get("street_key") or "").strip()

            deal = (row.get("deal_type") or "").strip().lower() or default_deal_type
            area = parse_num(row.get("area_m2"))
            price = parse_num(row.get("price_rub"))
            our_price = parse_num(row.get("our_best_price_rub"))
            pos = parse_num(row.get("position_global"))

            out.append(
                Listing(
                    competitor=competitor,
                    competitor_title=competitor_title,
                    deal_type=deal,
                    address=address,
                    area_m2=area,
                    price_rub=price,
                    result=(row.get("result") or "").strip(),
                    price_alert=(row.get("price_alert") or "").strip(),
                    position_global=pos,
                    competitor_link=(row.get("competitor_link") or "").strip(),
                    our_link=(row.get("our_best_link") or "").strip(),
                    our_best_price=our_price,
                    district=district,
                    district_norm=district_norm,
                    street_key=street_key,
                    address_key=build_address_key(address),
                )
            )
    return out


def find_matching_object(pool: List[UnionObject], lst: Listing) -> Optional[UnionObject]:
    if not pool:
        return None

    def district_compatible(obj: UnionObject) -> bool:
        if is_unknown_district(lst.district_norm):
            return True
        known = {x.district_norm for x in obj.listings.values() if not is_unknown_district(x.district_norm)}
        if not known:
            return True
        return lst.district_norm in known

    # 1) точное/близкое совпадение по площади
    best = None
    best_diff = 10**9
    for obj in pool:
        if not district_compatible(obj):
            continue
        if not isinstance(lst.area_m2, (int, float)) or not isinstance(obj.area_ref, (int, float)):
            continue
        d = abs(float(lst.area_m2) - float(obj.area_ref))
        if d <= AREA_TOL and d < best_diff:
            best = obj
            best_diff = d
    if best:
        return best

    # 1b) мягкий матч для единичного адреса (чтобы поймать "почти тот же объект")
    if (
        len(pool) == 1
        and district_compatible(pool[0])
        and isinstance(lst.area_m2, (int, float))
        and isinstance(pool[0].area_ref, (int, float))
    ):
        d = abs(float(lst.area_m2) - float(pool[0].area_ref))
        if d <= 15.0:
            return pool[0]

    # 2) если площадь отсутствует и объект в группе один — объединяем
    if lst.area_m2 is None and len(pool) == 1 and district_compatible(pool[0]):
        return pool[0]

    return None


def build_union_objects(listings: List[Listing]) -> List[UnionObject]:
    groups: Dict[str, List[UnionObject]] = {}

    for lst in listings:
        key = lst.address_key
        if key not in groups:
            groups[key] = []

        pool = groups[key]
        obj = find_matching_object(pool, lst)
        if obj is None:
            obj = UnionObject(address_key=key, area_ref=lst.area_m2 if isinstance(lst.area_m2, (int, float)) else None)
            pool.append(obj)
        obj.add_listing(lst)

    out = []
    for items in groups.values():
        out.extend(items)
    return out


def pick_display_address(obj: UnionObject) -> str:
    order = ["knru", "nordwest", "rest2rent"]
    for c in order:
        if c in obj.listings and obj.listings[c].address:
            return obj.listings[c].address
    for v in obj.listings.values():
        if v.address:
            return v.address
    return ""


def pick_display_district(obj: UnionObject) -> str:
    order = ["knru", "nordwest", "rest2rent"]
    for c in order:
        x = obj.listings.get(c)
        if x and x.district:
            return x.district
    return ""


def pick_min_position(obj: UnionObject) -> float:
    vals = [v.position_global for v in obj.listings.values() if isinstance(v.position_global, (int, float))]
    return min(vals) if vals else 10**12


def collect_diffs(obj: UnionObject) -> str:
    parts = []
    labels = {"knru": "KNRU", "nordwest": "Северо-Запад", "rest2rent": "Rest2Rent"}

    deal_vals = []
    for k, v in obj.listings.items():
        d = (v.deal_type or "").strip().lower()
        if d:
            deal_vals.append((labels.get(k, k), d))
    if len({x[1] for x in deal_vals}) >= 2:
        txt = " / ".join(f"{nm}: {dv}" for nm, dv in deal_vals)
        parts.append(f"Тип сделки отличается ({txt})")

    area_vals = []
    for k, v in obj.listings.items():
        if isinstance(v.area_m2, (int, float)):
            area_vals.append((labels.get(k, k), float(v.area_m2)))
    if len(area_vals) >= 2:
        a_numbers = [x[1] for x in area_vals]
        if max(a_numbers) - min(a_numbers) > AREA_TOL:
            txt = " / ".join(f"{nm}: {format_area(val)} м2" for nm, val in area_vals)
            parts.append(f"Площадь отличается ({txt})")

    price_vals = []
    for k, v in obj.listings.items():
        if isinstance(v.price_rub, (int, float)):
            price_vals.append((labels.get(k, k), float(v.price_rub)))
    if len(price_vals) >= 2:
        p_numbers = [x[1] for x in price_vals]
        if max(p_numbers) != min(p_numbers):
            txt = " / ".join(f"{nm}: {format_money(val)}" for nm, val in price_vals)
            parts.append(f"Цена отличается ({txt})")

    return " | ".join(parts)


def make_union_dataframe(objects: List[UnionObject]) -> pd.DataFrame:
    rows = []
    for obj in objects:
        kn = obj.listings.get("knru")
        nw = obj.listings.get("nordwest")
        rr = obj.listings.get("rest2rent")

        present = [x for x in (kn, nw, rr) if x is not None]
        presence_count = len(present)
        ours_missing_all = bool(present) and all((x.result or "").strip() == "Нет у нас" for x in present)
        red_flag = presence_count > 2 and ours_missing_all

        rows.append(
            {
                "Район": pick_display_district(obj),
                "Адрес": pick_display_address(obj),
                "Сделка": (kn or nw or rr).deal_type if (kn or nw or rr) else "",
                "KNRU": "Да" if kn else "",
                "Северо-Запад": "Да" if nw else "",
                "Rest2Rent": "Да" if rr else "",
                "Площадь KNRU": format_area(kn.area_m2) if kn else "",
                "Площадь СЗ": format_area(nw.area_m2) if nw else "",
                "Площадь R2R": format_area(rr.area_m2) if rr else "",
                "Цена KNRU": format_money(kn.price_rub) if kn else "",
                "Цена СЗ": format_money(nw.price_rub) if nw else "",
                "Цена R2R": format_money(rr.price_rub) if rr else "",
                "Вывод KNRU": kn.result if kn else "",
                "Вывод СЗ": nw.result if nw else "",
                "Вывод R2R": rr.result if rr else "",
                "Расхождения": collect_diffs(obj),
                "Ссылка KNRU": kn.competitor_link if kn else "",
                "Ссылка СЗ": nw.competitor_link if nw else "",
                "Ссылка R2R": rr.competitor_link if rr else "",
                "_presence_count": presence_count,
                "_sort_pos": pick_min_position(obj),
                "_district_sort": normalize_district(pick_display_district(obj)),
                "_street_sort": robot.norm_text(pick_display_address(obj)),
                "_red_flag": red_flag,
            }
        )

    df = pd.DataFrame(rows)
    if df.empty:
        return df

    # Сначала пересечения (3 -> 2 -> 1), затем район и улица.
    df = df.sort_values(by=["_presence_count", "_district_sort", "_street_sort", "_sort_pos"], ascending=[False, True, True, True])
    return df


def make_stats_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame([{"Показатель": "Нет данных", "Значение": 0}])

    kn = df["KNRU"].eq("Да")
    nw = df["Северо-Запад"].eq("Да")
    rr = df["Rest2Rent"].eq("Да")
    diff = df["Расхождения"].fillna("").str.len() > 0
    red = df["_red_flag"].fillna(False).astype(bool)

    rows = [
        {"Показатель": "Уникальных объединённых объектов", "Значение": int(len(df))},
        {"Показатель": "Есть у всех 3 конкурентов", "Значение": int((kn & nw & rr).sum())},
        {"Показатель": "Только KNRU", "Значение": int((kn & ~nw & ~rr).sum())},
        {"Показатель": "Только Северо-Запад", "Значение": int((~kn & nw & ~rr).sum())},
        {"Показатель": "Только Rest2Rent", "Значение": int((~kn & ~nw & rr).sum())},
        {"Показатель": "KNRU + Северо-Запад", "Значение": int((kn & nw & ~rr).sum())},
        {"Показатель": "KNRU + Rest2Rent", "Значение": int((kn & ~nw & rr).sum())},
        {"Показатель": "Северо-Запад + Rest2Rent", "Значение": int((~kn & nw & rr).sum())},
        {"Показатель": "Объекты с расхождениями значений", "Значение": int(diff.sum())},
        {"Показатель": "Есть у 3 конкурентов, у нас нет (красные)", "Значение": int(red.sum())},
    ]

    return pd.DataFrame(rows)


def parse_args():
    parser = argparse.ArgumentParser(description="Сводный отчёт по KNRU/Nordwest/Rest2Rent")
    parser.add_argument(
        "--knru",
        default=str(BASE_DIR / f"compare_report_{TODAY}.csv"),
        help="Путь к CSV KNRU",
    )
    parser.add_argument(
        "--nordwest",
        default=str(BASE_DIR / f"compare_report_nordwest_{TODAY}.csv"),
        help="Путь к CSV Nordwest",
    )
    parser.add_argument(
        "--rest2rent",
        default=str(BASE_DIR / f"compare_report_rest2rent_{TODAY}.csv"),
        help="Путь к CSV Rest2Rent",
    )
    parser.add_argument(
        "--output",
        default=str(BASE_DIR / f"compare_report_union_{TODAY}.xlsx"),
        help="Путь к выходному xlsx",
    )
    parser.add_argument(
        "--district-geocode",
        default="true",
        help="Доопределять пустые районы через Nominatim: true/false",
    )
    parser.add_argument(
        "--district-geocode-limit",
        type=int,
        default=500,
        help="Лимит geocode-запросов за запуск",
    )
    parser.add_argument(
        "--district-geocode-delay-sec",
        type=float,
        default=1.1,
        help="Пауза между geocode-запросами (сек)",
    )
    parser.add_argument(
        "--district-geocode-timeout-sec",
        type=int,
        default=8,
        help="Таймаут одного geocode-запроса (сек)",
    )
    parser.add_argument(
        "--district-cache",
        default=str(BASE_DIR / "district_cache_nominatim.json"),
        help="Файл кэша геокодинга районов",
    )
    return parser.parse_args()


def parse_bool(value: str | None, default: bool = True) -> bool:
    if value is None:
        return default
    s = str(value).strip().lower()
    if s in {"1", "true", "yes", "y", "on"}:
        return True
    if s in {"0", "false", "no", "n", "off"}:
        return False
    return default


def main():
    args = parse_args()
    knru_path = to_abs_path(args.knru)
    nordwest_path = to_abs_path(args.nordwest)
    rest_path = to_abs_path(args.rest2rent)
    out_path = to_abs_path(args.output)

    all_listings = []
    all_listings.extend(read_competitor_csv(knru_path, "knru", "KNRU", default_deal_type="sale"))
    all_listings.extend(read_competitor_csv(nordwest_path, "nordwest", "Северо-Запад", default_deal_type="sale"))
    all_listings.extend(read_competitor_csv(rest_path, "rest2rent", "Rest2Rent", default_deal_type="sale"))

    enrich_missing_districts(
        all_listings,
        use_geocode=parse_bool(args.district_geocode, True),
        geocode_limit=max(0, int(args.district_geocode_limit)),
        geocode_delay_sec=max(0.0, float(args.district_geocode_delay_sec)),
        geocode_timeout_sec=max(2, int(args.district_geocode_timeout_sec)),
        geocode_cache_path=to_abs_path(args.district_cache),
    )
    objects = build_union_objects(all_listings)
    union_df_full = make_union_dataframe(objects)
    stats_df = make_stats_dataframe(union_df_full)
    union_df = union_df_full.drop(
        columns=["_presence_count", "_sort_pos", "_district_sort", "_street_sort", "_red_flag"],
        errors="ignore",
    )

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        union_df.to_excel(writer, sheet_name="Сводный", index=False)
        stats_df.to_excel(writer, sheet_name="Статистика", index=False)

        # Красим строку в "Сводный", если объект есть у 3 конкурентов и везде "Нет у нас".
        wb = writer.book
        ws = wb["Сводный"]
        red_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
        n_cols = union_df.shape[1]
        for i, is_red in enumerate(union_df_full["_red_flag"].fillna(False).tolist(), start=2):
            if not is_red:
                continue
            for col in range(1, n_cols + 1):
                ws.cell(row=i, column=col).fill = red_fill

    print(f"Saved: {out_path}")
    print(stats_df.to_string(index=False))


if __name__ == "__main__":
    main()
